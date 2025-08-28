"""
Professional Office documents to PDF conversion service.
Supports Excel, PowerPoint, and other Office formats with high-quality output.
"""

import asyncio
import io
import logging
import subprocess
from datetime import datetime
from typing import Any, Dict, Optional, List
from pathlib import Path
import base64
import tempfile
import os

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from pptx import Presentation
    PPTX_AVAILABLE = True
except ImportError:
    PPTX_AVAILABLE = False

# Windows COM support
try:
    import comtypes.client
    COMTYPES_AVAILABLE = True
except ImportError:
    COMTYPES_AVAILABLE = False

from ..core.base import (
    BaseConverter, 
    ConversionFormat, 
    ConversionRequest, 
    ConversionResult,
    ConversionStatus,
    ProcessingError,
    ValidationError
)

logger = logging.getLogger(__name__)


class OfficeToPdfConverter(BaseConverter):
    """
    Professional Office documents to PDF conversion service.
    
    Supported formats:
    - Excel (.xlsx, .xls) 
    - PowerPoint (.pptx, .ppt)
    - Word (.docx, .doc) - via separate service
    - OpenDocument formats (.ods, .odp, .odt)
    
    Features:
    - High-quality PDF output
    - Preserves formatting and layout
    - Multiple conversion engines
    - Batch processing support
    - Custom page settings
    - Print area optimization
    - Chart and image handling
    """
    
    def __init__(self):
        super().__init__(
            name="office_to_pdf",
            supported_formats=[ConversionFormat.PDF]
        )
        
        # Determine available conversion methods
        self.conversion_methods = []
        
        if COMTYPES_AVAILABLE and os.name == 'nt':  # Windows only
            self.conversion_methods.append("office_com")
        
        # LibreOffice as universal fallback
        if self._check_libreoffice():
            self.conversion_methods.append("libreoffice")
        
        # Python libraries for specific formats
        if OPENPYXL_AVAILABLE:
            self.conversion_methods.append("openpyxl_html")
        
        if not self.conversion_methods:
            logger.warning("No Office to PDF conversion methods available")
        
        # Supported input formats
        self.supported_input_formats = {
            '.xlsx': 'excel',
            '.xls': 'excel', 
            '.xlsm': 'excel',
            '.pptx': 'powerpoint',
            '.ppt': 'powerpoint',
            '.pptm': 'powerpoint',
            '.ods': 'calc',
            '.odp': 'impress',
            '.odt': 'writer'
        }
    
    def _check_libreoffice(self) -> bool:
        """Check if LibreOffice is available."""
        try:
            result = subprocess.run(
                ['libreoffice', '--version'], 
                capture_output=True, 
                text=True, 
                timeout=10
            )
            return result.returncode == 0
        except Exception:
            return False
    
    def validate_request(self, request: ConversionRequest) -> bool:
        """Validate Office to PDF conversion request."""
        if not request.content or not request.content.strip():
            return False
        
        if request.output_format != ConversionFormat.PDF:
            return False
        
        # Check if format is supported
        options = request.options or {}
        input_format = options.get('input_format', '').lower()
        
        if input_format and f'.{input_format}' not in self.supported_input_formats:
            return False
        
        return True
    
    async def convert(self, request: ConversionRequest) -> ConversionResult:
        """
        Convert Office document to PDF.
        
        Args:
            request: Conversion request with Office document content
            
        Returns:
            ConversionResult with PDF data
        """
        job_id = f"office2pdf_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{id(request)}"
        
        try:
            logger.info(f"Starting Office to PDF conversion: {job_id}")
            
            # Validate request
            if not self.validate_request(request):
                raise ValidationError("Invalid Office to PDF conversion request", job_id)
            
            # Extract options and determine format
            options = request.options or {}
            input_format = await self._detect_format(request.content, options)
            conversion_method = options.get('method', 'auto')
            
            # Convert Office document to PDF
            pdf_data = await self._convert_office_to_pdf(
                request.content, input_format, options, conversion_method
            )
            
            # Create result
            result = ConversionResult(
                job_id=job_id,
                status=ConversionStatus.COMPLETED,
                output_data=pdf_data,
                output_filename=f"{job_id}.pdf",
                metadata={
                    "original_size": len(request.content),
                    "pdf_size": len(pdf_data),
                    "input_format": input_format,
                    "conversion_method": conversion_method,
                    "options_used": options
                },
                created_at=datetime.utcnow(),
                completed_at=datetime.utcnow()
            )
            
            logger.info(f"Office to PDF conversion completed: {job_id}")
            return result
            
        except Exception as e:
            logger.error(f"Office to PDF conversion failed: {job_id} - {str(e)}")
            return ConversionResult(
                job_id=job_id,
                status=ConversionStatus.FAILED,
                error_message=str(e),
                created_at=datetime.utcnow(),
                completed_at=datetime.utcnow()
            )
    
    async def _detect_format(self, content: str, options: Dict) -> str:
        """Detect Office document format."""
        # Check explicit format in options
        if options.get('input_format'):
            return options['input_format']
        
        # Try to detect from content
        if content.startswith('data:'):
            # Extract from data URL
            header = content.split(',')[0]
            if 'excel' in header or 'spreadsheet' in header:
                return 'xlsx'
            elif 'powerpoint' in header or 'presentation' in header:
                return 'pptx'
        
        # Check if it's a file path
        if Path(content).exists():
            suffix = Path(content).suffix.lower()
            if suffix in self.supported_input_formats:
                return suffix[1:]  # Remove dot
        
        # Default to Excel if uncertain
        return 'xlsx'
    
    async def _convert_office_to_pdf(self, content: str, input_format: str, options: Dict, method: str) -> bytes:
        """Convert Office document to PDF using specified method."""
        
        # Determine conversion method
        if method == 'auto':
            method = self._select_best_method(input_format, options)
        
        if method not in self.conversion_methods:
            raise ProcessingError(f"Conversion method '{method}' not available")
        
        # Create temporary files
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_dir_path = Path(temp_dir)
            input_ext = f".{input_format}" if not input_format.startswith('.') else input_format
            input_path = temp_dir_path / f"input{input_ext}"
            pdf_path = temp_dir_path / "output.pdf"
            
            # Write input content to file
            await self._write_office_content(content, input_path)
            
            # Convert based on method
            if method == "office_com":
                await self._convert_with_office_com(input_path, pdf_path, input_format, options)
            elif method == "libreoffice":
                await self._convert_with_libreoffice(input_path, pdf_path, options)
            elif method == "openpyxl_html" and input_format in ['xlsx', 'xls']:
                await self._convert_excel_via_html(input_path, pdf_path, options)
            else:
                raise ProcessingError(f"Unknown conversion method: {method}")
            
            # Read PDF data
            if not pdf_path.exists():
                raise ProcessingError("PDF conversion failed - output file not created")
            
            return pdf_path.read_bytes()
    
    async def _write_office_content(self, content: str, output_path: Path):
        """Write Office document content to file."""
        if content.startswith('data:'):
            # Base64 encoded content
            header, data = content.split(',', 1)
            content_bytes = base64.b64decode(data)
        elif len(content) > 1000 and not content.startswith('<'):
            # Likely binary content
            content_bytes = content.encode('latin1')
        else:
            # File path
            if Path(content).exists():
                content_bytes = Path(content).read_bytes()
            else:
                raise ProcessingError("Invalid Office document content")
        
        output_path.write_bytes(content_bytes)
    
    async def _convert_with_office_com(self, input_path: Path, pdf_path: Path, input_format: str, options: Dict):
        """Convert using Office COM interface (Windows only)."""
        loop = asyncio.get_event_loop()
        
        def convert():
            if input_format in ['xlsx', 'xls', 'xlsm']:
                self._convert_excel_com(input_path, pdf_path, options)
            elif input_format in ['pptx', 'ppt', 'pptm']:
                self._convert_powerpoint_com(input_path, pdf_path, options)
            else:
                raise ProcessingError(f"COM conversion not supported for {input_format}")
        
        await loop.run_in_executor(None, convert)
    
    def _convert_excel_com(self, input_path: Path, pdf_path: Path, options: Dict):
        """Convert Excel using COM interface."""
        excel = comtypes.client.CreateObject('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            workbook = excel.Workbooks.Open(str(input_path.absolute()))
            
            # Configure print settings
            for worksheet in workbook.Worksheets:
                if options.get('fit_to_page'):
                    worksheet.PageSetup.Zoom = False
                    worksheet.PageSetup.FitToPagesWide = options.get('pages_wide', 1)
                    worksheet.PageSetup.FitToPagesTall = options.get('pages_tall', False)
                
                if options.get('orientation'):
                    # 1 = Portrait, 2 = Landscape
                    worksheet.PageSetup.Orientation = 2 if options['orientation'] == 'landscape' else 1
                
                # Set print area if specified
                if options.get('print_area'):
                    worksheet.PageSetup.PrintArea = options['print_area']
            
            # Export as PDF
            workbook.ExportAsFixedFormat(
                Type=0,  # PDF format
                Filename=str(pdf_path.absolute()),
                Quality=0,  # Minimum size
                IncludeDocProps=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            workbook.Close(SaveChanges=False)
        finally:
            excel.Quit()
    
    def _convert_powerpoint_com(self, input_path: Path, pdf_path: Path, options: Dict):
        """Convert PowerPoint using COM interface."""
        powerpoint = comtypes.client.CreateObject('PowerPoint.Application')
        powerpoint.Visible = 0  # Hidden
        
        try:
            presentation = powerpoint.Presentations.Open(
                str(input_path.absolute()),
                ReadOnly=True,
                Untitled=True,
                WithWindow=False
            )
            
            # Export as PDF
            presentation.ExportAsFixedFormat(
                Path=str(pdf_path.absolute()),
                FixedFormatType=2,  # PDF
                Intent=1,  # Print intent
                FrameSlides=0,  # Don't frame slides
                HandoutOrder=1,  # Horizontal order
                OutputType=0,  # All slides
                PrintHiddenSlides=0,  # Don't print hidden
                PrintRange=None,
                RangeType=0,  # All slides
                SlideShowName='',
                IncludeDocProps=True,
                KeepIRMSettings=True,
                DocStructureTags=True,
                BitmapMissingFonts=True,
                UseDocumentOrder=True
            )
            
            presentation.Close()
        finally:
            powerpoint.Quit()
    
    async def _convert_with_libreoffice(self, input_path: Path, pdf_path: Path, options: Dict):
        """Convert using LibreOffice."""
        cmd = [
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', str(input_path.parent),
            str(input_path)
        ]
        
        # Add LibreOffice-specific options
        if options.get('export_filter'):
            cmd.extend(['--infilter', options['export_filter']])
        
        process = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        
        stdout, stderr = await process.communicate()
        
        if process.returncode != 0:
            raise ProcessingError(f"LibreOffice conversion failed: {stderr.decode()}")
        
        # LibreOffice creates PDF with same name as input
        generated_pdf = input_path.parent / f"{input_path.stem}.pdf"
        if generated_pdf.exists():
            generated_pdf.rename(pdf_path)
        else:
            raise ProcessingError("LibreOffice did not generate expected PDF file")
    
    async def _convert_excel_via_html(self, input_path: Path, pdf_path: Path, options: Dict):
        """Convert Excel to PDF via HTML (for better table formatting)."""
        # Load Excel file
        workbook = load_workbook(input_path, data_only=True)
        
        # Convert to HTML
        html_content = await self._excel_to_html(workbook, options)
        
        # Use HTML to PDF converter
        from .html_pdf import HtmlToPdfConverter
        html_converter = HtmlToPdfConverter()
        
        html_request = type('Request', (), {
            'content': html_content,
            'output_format': ConversionFormat.PDF,
            'options': {
                'engine': 'playwright' if 'playwright' in html_converter.engines else html_converter.engines[0],
                'page_size': options.get('page_size', 'A4'),
                'landscape': options.get('orientation') == 'landscape',
                'print_background': True,
                'margins': options.get('margins', {
                    'top': '1cm',
                    'bottom': '1cm', 
                    'left': '1cm',
                    'right': '1cm'
                })
            }
        })()
        
        result = await html_converter.convert(html_request)
        
        if result.status == ConversionStatus.COMPLETED:
            pdf_path.write_bytes(result.output_data)
        else:
            raise ProcessingError(f"HTML to PDF conversion failed: {result.error_message}")
    
    async def _excel_to_html(self, workbook: Workbook, options: Dict) -> str:
        """Convert Excel workbook to HTML."""
        html_parts = ['<!DOCTYPE html>', '<html><head>']
        html_parts.append('<meta charset="utf-8">')
        html_parts.append('<style>')
        html_parts.append(self._get_excel_html_styles())
        html_parts.append('</style>')
        html_parts.append('</head><body>')
        
        # Process worksheets
        worksheets = options.get('worksheets', workbook.sheetnames)
        if isinstance(worksheets, str):
            worksheets = [worksheets]
        
        for sheet_name in worksheets:
            if sheet_name not in workbook.sheetnames:
                continue
                
            worksheet = workbook[sheet_name]
            
            if len(worksheets) > 1:
                html_parts.append(f'<h2>{sheet_name}</h2>')
            
            html_parts.append('<table>')
            
            # Get used range
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            for row_num in range(1, max_row + 1):
                html_parts.append('<tr>')
                
                for col_num in range(1, max_col + 1):
                    cell = worksheet.cell(row_num, col_num)
                    cell_value = cell.value if cell.value is not None else ''
                    
                    # Apply basic formatting
                    style_attrs = []
                    if cell.font and cell.font.bold:
                        style_attrs.append('font-weight: bold')
                    if cell.font and cell.font.italic:
                        style_attrs.append('font-style: italic')
                    
                    style_str = f' style="{"; ".join(style_attrs)}"' if style_attrs else ''
                    
                    html_parts.append(f'<td{style_str}>{cell_value}</td>')
                
                html_parts.append('</tr>')
            
            html_parts.append('</table>')
            
            if len(worksheets) > 1:
                html_parts.append('<div style="page-break-after: always;"></div>')
        
        html_parts.append('</body></html>')
        return '\\n'.join(html_parts)
    
    def _get_excel_html_styles(self) -> str:
        """Get CSS styles for Excel HTML conversion."""
        return """
        body {
            font-family: Calibri, Arial, sans-serif;
            font-size: 11pt;
            margin: 0.5in;
        }
        h2 {
            font-size: 14pt;
            margin: 0 0 10pt 0;
            font-weight: bold;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin-bottom: 20pt;
        }
        td {
            border: 1px solid #ccc;
            padding: 2pt 4pt;
            text-align: left;
            vertical-align: top;
        }
        tr:first-child td {
            background-color: #f0f0f0;
            font-weight: bold;
        }
        """
    
    def _select_best_method(self, input_format: str, options: Dict) -> str:
        """Select the best available conversion method."""
        # Preference order based on quality and format support
        if input_format in ['xlsx', 'xls', 'xlsm']:
            preferred_order = ["office_com", "openpyxl_html", "libreoffice"]
        elif input_format in ['pptx', 'ppt', 'pptm']:
            preferred_order = ["office_com", "libreoffice"]
        else:
            preferred_order = ["libreoffice", "office_com"]
        
        for method in preferred_order:
            if method in self.conversion_methods:
                return method
        
        raise ProcessingError("No conversion methods available for this format")
    
    def get_capabilities(self) -> Dict[str, Any]:
        """Get converter capabilities."""
        return {
            "name": self.name,
            "description": "Professional Office documents to PDF conversion service",
            "supported_formats": [fmt.value for fmt in self.supported_formats],
            "supported_input_formats": list(self.supported_input_formats.keys()),
            "available_methods": self.conversion_methods,
            "features": [
                "High-quality PDF output",
                "Preserves formatting and layout", 
                "Multiple conversion engines",
                "Batch processing support",
                "Custom page settings",
                "Print area optimization",
                "Chart and image handling"
            ],
            "supported_options": [
                "method",
                "input_format",
                "page_size",
                "orientation",
                "margins",
                "fit_to_page",
                "print_area",
                "worksheets",
                "export_filter"
            ],
            "example_request": {
                "content": "Excel/PowerPoint file content or path",
                "output_format": "pdf",
                "options": {
                    "input_format": "xlsx",
                    "method": "auto",
                    "orientation": "landscape",
                    "fit_to_page": True
                }
            }
        }